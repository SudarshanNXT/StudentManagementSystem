import tkinter as tk
import tkinter.ttk as ttk
import tkinter.messagebox as tkMessageBox
from tkcalendar import DateEntry
from datetime import datetime
import sqlite3
from sqlite3 import Error
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF

USERNAME = "admin"
PASSWORD = "password"

def Database():
    conn = None
    try:
        conn = sqlite3.connect("student.dbms")
        if conn:
            print("Connected to SQLite database")
            cursor = conn.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS STUD_REGISTRATION (
                    STU_ID TEXT PRIMARY KEY,
                    STU_NAME TEXT,
                    STU_CONTACT INTEGER,
                    STU_EMAIL TEXT,
                    STU_ROLLNO INTEGER UNIQUE,
                    STU_BRANCH TEXT,
                    STU_DOB DATE,
                    STU_ADDRESS TEXT,
                    FEES_PAID TEXT,
                    YEAR TEXT
                )
            """)
            conn.commit()
            return conn, cursor
    except Error as e:
        print("Error:", e)
        if conn:
            conn.close()
        return None, None

def DisplayForm():
    global display_screen, date_label
    display_screen = tk.Tk()
    display_screen.geometry("900x400")
    display_screen.title("Student Management System")
    display_screen.configure(bg="light green")
    display_screen.iconbitmap('C:/Users/white/OneDrive/Desktop/sudarshan/SMS/icon.ico')

    LoginWindow()
    date_label = tk.Label(display_screen, text="", font=("Bold", 10), bg="light green", fg="black")
    date_label.pack(anchor=tk.NE, padx=10, pady=10)
    update_datetime()

def update_datetime():
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    current_date = now.strftime("%d-%m-%Y")
    datetime_str = f"{current_date} | {current_time}"
    date_label.config(text=datetime_str)
    display_screen.after(1000, update_datetime)

def LoginWindow():
    global login_screen, username_entry, password_entry

    def Login():
        username = username_entry.get()
        password = password_entry.get()

        if username == USERNAME and password == PASSWORD:
            login_screen.destroy()
            MainApplication()
            tkMessageBox.showinfo("Login Successful", f"Good Day {username}!")
        else:
            tkMessageBox.showerror("Login Failed", "Invalid username or password")

    login_screen = tk.Toplevel(display_screen)
    login_screen.title("Login")
    login_screen.configure(bg="light green")
    login_screen.iconbitmap('C:/Users/white/OneDrive/Desktop/sudarshan/SMS/icon.ico')

    tk.Label(login_screen, text="ADMIN PANEL", font=("Arial", 16, "bold"), bg="light green", fg="black").pack(pady=(10, 20))

    tk.Label(login_screen, text="Username", bg="light green", fg="black").pack()
    username_entry = tk.Entry(login_screen)
    username_entry.pack()

    tk.Label(login_screen, text="Password", bg="light green", fg="black").pack()
    password_entry = tk.Entry(login_screen, show="*")
    password_entry.pack()
    tk.Label(login_screen, text="Contact Administration for Login Details", bg="light green", fg="black").pack(pady=(10, 0))

    tk.Button(login_screen, text="Login", command=Login, bg="black", fg="white").pack(pady=20)

def MainApplication():
    global tree, SEARCH, name, contact, email, rollno, course, dob_entry, address, fees_paid_var, year_var
    SEARCH = tk.StringVar()
    name = tk.StringVar()
    contact = tk.StringVar()
    email = tk.StringVar()
    rollno = tk.StringVar()
    course = tk.StringVar()
    dob = tk.StringVar()
    address = tk.StringVar()
    fees_paid_var = tk.StringVar()
    year_var = tk.StringVar()

    TopViewForm = tk.Frame(display_screen, width=600, bd=1, relief=tk.SOLID, bg="light green")
    TopViewForm.pack(side=tk.TOP, fill=tk.X)
    LFrom = tk.Frame(display_screen, width="350", bg="light green")
    LFrom.pack(side=tk.LEFT, fill=tk.Y)
    LeftViewForm = tk.Frame(display_screen, width=500, bg="gray")
    LeftViewForm.pack(side=tk.LEFT, fill=tk.Y)
    MidViewForm = tk.Frame(display_screen, width=600, bg="light green")
    MidViewForm.pack(side=tk.RIGHT)

    lbl_text = tk.Label(TopViewForm, text="Student Management System", font=('verdana', 18), width=600, bg="light green", fg="black")
    lbl_text.pack(fill=tk.X)

    tk.Label(LFrom, text="Name", font=("Arial", 12), bg="light green", fg="black").pack(side=tk.TOP)
    tk.Entry(LFrom, font=("Arial", 10, "bold"), textvariable=name).pack(side=tk.TOP, padx=10, fill=tk.X)
    tk.Label(LFrom, text="Contact", font=("Arial", 12), bg="light green", fg="black").pack(side=tk.TOP)
    tk.Entry(LFrom, font=("Arial", 10, "bold"), textvariable=contact).pack(side=tk.TOP, padx=10, fill=tk.X)
    tk.Label(LFrom, text="Email", font=("Arial", 12), bg="light green", fg="black").pack(side=tk.TOP)
    tk.Entry(LFrom, font=("Arial", 10, "bold"), textvariable=email).pack(side=tk.TOP, padx=10, fill=tk.X)
    tk.Label(LFrom, text="Roll.No", font=("Arial", 12), bg="light green", fg="black").pack(side=tk.TOP)
    tk.Entry(LFrom, font=("Arial", 10, "bold"), textvariable=rollno).pack(side=tk.TOP, padx=10, fill=tk.X)
    tk.Label(LFrom, text="Course", font=("Arial", 12), bg="light green", fg="black").pack(side=tk.TOP)
    tk.Entry(LFrom, font=("Arial", 10, "bold"), textvariable=course).pack(side=tk.TOP, padx=10, fill=tk.X)
    tk.Label(LFrom, text="Date of Birth", font=("Arial", 12), bg="light green", fg="black").pack(side=tk.TOP)
    dob_entry = DateEntry(LFrom, width=12, background='darkyellow', foreground='white', borderwidth=2, year=1990)
    dob_entry.pack(side=tk.TOP, padx=10, fill=tk.X)
    tk.Label(LFrom, text="Address", font=("Arial", 12), bg="light green", fg="black").pack(side=tk.TOP)
    tk.Entry(LFrom, font=("Arial", 10, "bold"), textvariable=address).pack(side=tk.TOP, padx=10, fill=tk.X)
    tk.Label(LFrom, text="Fees Paid", font=("Arial", 12), bg="light green", fg="black").pack(side=tk.TOP)
    ttk.Combobox(LFrom, font=("Arial", 10, "bold"), textvariable=fees_paid_var, values=["Yes", "No"]).pack(side=tk.TOP, padx=10, fill=tk.X)

    tk.Label(LFrom, text="Year", font=("Arial", 12), bg="light green", fg="black").pack(side=tk.TOP)
    ttk.Combobox(LFrom, font=("Arial", 10, "bold"), textvariable=year_var, values=["FY", "SY", "TE", "BE"]).pack(side=tk.TOP, padx=10, fill=tk.X)

    tk.Button(LFrom, text="Submit", font=("Arial", 10, "bold"), command=register, bg="black", fg="white").pack(side=tk.TOP, padx=10, pady=5, fill=tk.X)
    tk.Button(LFrom, text="Reset", font=("Arial", 10, "bold"), command=Reset, bg="black", fg="white").pack(side=tk.TOP, padx=10, pady=5, fill=tk.X)

    lbl_txtsearch = tk.Label(LeftViewForm, text="Enter ROLL.NO", font=('verdana', 10), bg="gray", fg="black")
    lbl_txtsearch.pack()
    search = tk.Entry(LeftViewForm, textvariable=SEARCH, font=('verdana', 15), width=10)
    search.pack(side=tk.TOP, padx=10, fill=tk.X)
    tk.Button(LeftViewForm, text="Search", command=SearchRecord, bg="black", fg="white").pack(side=tk.TOP, padx=10, pady=10, fill=tk.X)
    tk.Button(LeftViewForm, text="View All", command=DisplayData, bg="black", fg="white").pack(side=tk.TOP, padx=10, pady=10, fill=tk.X)
    tk.Button(LeftViewForm, text="Update", command=update, bg="black", fg="white").pack(side=tk.TOP, padx=10, pady=10, fill=tk.X)
    tk.Button(LeftViewForm, text="Delete", command=Delete, bg="black", fg="white").pack(side=tk.TOP, padx=10, pady=10, fill=tk.X)
    tk.Button(LeftViewForm, text="Clear ALL", command=ClearAllRecords, bg="black", fg="white").pack(side=tk.TOP, padx=10, pady=10, fill=tk.X)

    # Add button to export to Excel and PDF
    tk.Button(LeftViewForm, text="Export to Excel", command=export_to_excel, bg="black", fg="white").pack(side=tk.TOP, padx=10, pady=10, fill=tk.X)
    tk.Button(LeftViewForm, text="Export to PDF", command=export_to_pdf, bg="black", fg="white").pack(side=tk.TOP, padx=10, pady=10, fill=tk.X)

    tk.Button(LeftViewForm, text="Help Center", command=show_help_message, bg="black", fg="white").pack(side=tk.TOP, padx=10, pady=10, fill=tk.X)

    scrollbarx = ttk.Scrollbar(MidViewForm, orient=tk.HORIZONTAL)
    scrollbary = ttk.Scrollbar(MidViewForm, orient=tk.VERTICAL)
    tree = ttk.Treeview(MidViewForm, columns=("Student Id", "Name", "Contact", "Email", "Rollno", "Course", "Date of Birth", "Address", "Fees Paid", "Year"),
                        selectmode="extended", height=100, yscrollcommand=scrollbary.set, xscrollcommand=scrollbarx.set)
    scrollbary.config(command=tree.yview)
    scrollbary.pack(side=tk.RIGHT, fill=tk.Y)
    scrollbarx.config(command=tree.xview)
    scrollbarx.pack(side=tk.BOTTOM, fill=tk.X)

    tree.heading('Student Id', text="Student Id", anchor=tk.W)
    tree.heading('Name', text="Name", anchor=tk.W)
    tree.heading('Contact', text="Contact", anchor=tk.W)
    tree.heading('Email', text="Email", anchor=tk.W)
    tree.heading('Rollno', text="Rollno", anchor=tk.W)
    tree.heading('Course', text="Course", anchor=tk.W)
    tree.heading('Date of Birth', text="Date of Birth", anchor=tk.W)
    tree.heading('Address', text="Address", anchor=tk.W)
    tree.heading('Fees Paid', text="Fees Paid", anchor=tk.W)
    tree.heading('Year', text="Year", anchor=tk.W)

    tree.column('#0', stretch=tk.NO, minwidth=0, width=0)
    tree.column('#1', stretch=tk.NO, minwidth=0, width=100)
    tree.column('#2', stretch=tk.NO, minwidth=0, width=150)
    tree.column('#3', stretch=tk.NO, minwidth=0, width=80)
    tree.pack()
    DisplayData()
    tree.bind("<<TreeviewSelect>>", on_select)  # Bind on_select to Treeview selection
    menubar = tk.Menu(display_screen)
    menubar.add_command(label="Logout❌", command=Logout)
    display_screen.config(menu=menubar)

def is_valid_contact(contact):
    return contact.isdigit()

def is_valid_rollno(rollno):
    return rollno.isdigit()

def register():
    name1 = name.get()
    con1 = contact.get()
    email1 = email.get()
    rol1 = rollno.get()
    course1 = course.get()
    dob1 = dob_entry.get()
    addr1 = address.get()
    fees_paid = fees_paid_var.get()
    year = year_var.get()

    if name1 == '' or con1 == '' or email1 == '' or rol1 == '' or course1 == '' or dob1 == '' or addr1 == '' or year == '':
        tkMessageBox.showinfo("Warning", "Fill all the fields!!!")
    elif not is_valid_contact(con1):
        tkMessageBox.showinfo("Warning", "Invalid contact number. Please enter numbers only.")
    elif not is_valid_rollno(rol1):
        tkMessageBox.showinfo("Warning", "Invalid roll number. Please enter numbers only.")
    else:
        try:
            conn, cursor = Database()
            cursor.execute("SELECT MAX(STU_ID) FROM STUD_REGISTRATION")
            last_stu_id = cursor.fetchone()[0]

            if last_stu_id:
                last_stu_id_number = int(last_stu_id[4:])
                next_stu_id_number = last_stu_id_number + 1
                next_stu_id = f"IOIT{next_stu_id_number:03}"
            else:
                next_stu_id = "IOIT100"

            cursor.execute("SELECT STU_ROLLNO FROM STUD_REGISTRATION WHERE STU_ROLLNO = ?", (rol1,))
            existing_roll = cursor.fetchone()

            if existing_roll:
                tkMessageBox.showerror("Error", "Roll number already exists.")
            else:
                cursor.execute('INSERT INTO STUD_REGISTRATION (STU_ID, STU_NAME, STU_CONTACT, STU_EMAIL, STU_ROLLNO, STU_BRANCH, STU_DOB, STU_ADDRESS, FEES_PAID, YEAR) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                               (next_stu_id, name1, con1, email1, rol1, course1, dob1, addr1, fees_paid, year))
                conn.commit()
                tkMessageBox.showinfo("Message", "Stored successfully")
                DisplayData()
        except Error as e:
            print("Error:", e)


def on_select(event):
    selected_item = tree.selection()
    if selected_item:
        item_data = tree.item(selected_item)['values']
        if item_data:
            name.set(item_data[1])
            contact.set(item_data[2])
            email.set(item_data[3])
            rollno.set(item_data[4])
            course.set(item_data[5])
            dob_entry.set_date(item_data[6])
            address.set(item_data[7])
            fees_paid_var.set(item_data[8])
            year_var.set(item_data[9])

def update():
    selected_item = tree.selection()
    if not selected_item:
        tkMessageBox.showwarning("Warning", "Select a record to update.")
        return

    name1 = name.get()
    con1 = contact.get()
    email1 = email.get()
    rol1 = rollno.get()
    course1 = course.get()
    dob1 = dob_entry.get()
    addr1 = address.get()
    fees_paid = fees_paid_var.get()
    year = year_var.get()

    if name1 == '' or con1 == '' or email1 == '' or rol1 == '' or course1 == '' or dob1 == '' or addr1 == '':
        tkMessageBox.showinfo("Warning", "Fill all the fields!!!")
        return

    try:
        conn, cursor = Database()
        item_id = selected_item[0]
        item_data = tree.item(item_id)['values']
        db_id = item_data[0]

        if rol1 != item_data[4]:
            cursor.execute("SELECT STU_ROLLNO FROM STUD_REGISTRATION WHERE STU_ROLLNO = ? AND STU_ID != ?", (rol1, db_id))
            existing_roll = cursor.fetchone()
            if existing_roll:
                tkMessageBox.showerror("Error", "Roll number already exists.")
                return

        cursor.execute('''UPDATE STUD_REGISTRATION 
                          SET STU_NAME=?, STU_CONTACT=?, STU_EMAIL=?, STU_ROLLNO=?, STU_BRANCH=?, 
                          STU_DOB=?, STU_ADDRESS=?, FEES_PAID=?, YEAR=? 
                          WHERE STU_ID=?''',
                       (name1, con1, email1, rol1, course1, dob1, addr1, fees_paid, year, db_id))
        conn.commit()
        tkMessageBox.showinfo("Message", "Record updated successfully.")
        DisplayData()
    except Error as e:
        print("Error:", e)


def Reset():
    tree.delete(*tree.get_children())
    DisplayData()
    SEARCH.set("")
    name.set("")
    contact.set("")
    email.set("")
    rollno.set("")
    course.set("")
    address.set("")
    fees_paid_var.set("")
    year_var.set("")
    dob_entry.set_date(1990)
    
def Delete():
    if not tree.selection():
        tkMessageBox.showwarning("Warning", "Select data to delete")
    else:
        result = tkMessageBox.askquestion('Confirm', 'Are you sure you want to delete this record?', icon="warning")
        if result == 'yes':
            curItem = tree.focus()
            contents = (tree.item(curItem))
            selecteditem = contents['values']
            tree.delete(curItem)
            try:
                conn, cursor = Database()
                cursor.execute("DELETE FROM STUD_REGISTRATION WHERE STU_ID = ?", (selecteditem[0],))
                conn.commit()
            except Error as e:
                print("Error:", e)

def ClearAllRecords():
    result = tkMessageBox.askquestion('Confirm', 'Are you sure you want to delete all records?', icon="warning")
    if result == 'yes':
        try:
            conn, cursor = Database()
            cursor.execute("DELETE FROM STUD_REGISTRATION")
            conn.commit()
            tkMessageBox.showinfo("Message", "All records deleted successfully.")
            DisplayData()
        except Error as e:
            print("Error:", e)

def SearchRecord():
    roll_number = SEARCH.get().strip()

    if roll_number != "":
        tree.delete(*tree.get_children())
        try:
            conn, cursor = Database()
            cursor.execute("SELECT * FROM STUD_REGISTRATION WHERE STU_ROLLNO LIKE ?", ('%' + roll_number + '%',))
            fetch = cursor.fetchall()
            if fetch:
                for data in fetch:
                    tree.insert('', 'end', values=data)
            else:
                tkMessageBox.showinfo("No Records", f"No records found for roll number: {roll_number}")
        except Error as e:
            print("Error:", e)

def DisplayData():
    tree.delete(*tree.get_children())
    try:
        conn, cursor = Database()
        cursor.execute("SELECT * FROM STUD_REGISTRATION")
        fetch = cursor.fetchall()
        for data in fetch:
            tree.insert('', 'end', values=data)
    except Error as e:
        print("Error:", e)

def Logout():
    global display_screen
    display_screen.destroy()
    DisplayForm()
    tkMessageBox.showinfo("Logout", "Thank you, Admin!", icon="info", fg="red")

def export_to_excel():
    try:
        conn, cursor = Database()
        cursor.execute("SELECT * FROM STUD_REGISTRATION")
        data = cursor.fetchall()
        df = pd.DataFrame(data, columns=["Student Id", "Name", "Contact", "Email", "Rollno", "Course", "Date of Birth", "Address", "Fees Paid", "Year"])

        # Create a new workbook
        wb = Workbook()
        ws = wb.active

        # Append dataframe to worksheet
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        # Save the workbook
        wb.save("student_data.xlsx")
        tkMessageBox.showinfo("Export Successful", "Data exported to 'student_data.xlsx' successfully.")
    except Error as e:
        print("Error:", e)

def export_to_pdf():
    try:
        conn, cursor = Database()
        cursor.execute("SELECT * FROM STUD_REGISTRATION")
        records = cursor.fetchall()

        if records:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            for record in records:
                pdf.cell(200, 10, f"Student Name: {record[1]}, Roll No: {record[4]}, Contact: {record[2]}", ln=True)
                pdf.cell(200, 10, f"Email: {record[3]}, Branch: {record[5]}", ln=True)
                pdf.cell(200, 10, f"DOB: {record[6]}, Address: {record[7]}, Fees Paid:{record[8]} ", ln=True)
                pdf.cell(200, 10, f"Year: {record[9]}", ln=True)
                pdf.cell(200, 10, "-" * 50, ln=True)
            file_path = 'student_data.pdf'
            pdf.output(file_path)
            tkMessageBox.showinfo("Export Successful", f"Data exported to '{file_path}' successfully.")
        else:
            tkMessageBox.showinfo("Export Failed", "No records found to export.")
    except Error as e:
        print("Error:", e)

def show_help_message():
    help_text = """
    This is the Student Management System Help Center.
    
    To get started:
    1. Use the fields on the left to enter student details.
    2. Click 'Submit' to save the student's information.
    3. Use the 'Search' box to find a student by roll number.
    4. Click 'View All' to see all student records.
    5. To update a record, select it from the table, edit, and click 'Update'.
    6. To delete a record, select it from the table and click 'Delete'.
    7. Use 'Export to Excel' or 'Export to PDF' to save student data.
    8. Logout from the menu when finished.
    
    For any assistance, please contact the administrator.
    📞 9000000012 // 📧 sudarshannxt@icloud.com


    Developed By SudarshanNXT®
    """
    tkMessageBox.showinfo("Help Center", help_text)
    
def Logout():
    global display_screen
    display_screen.destroy()
    DisplayForm()
    tkMessageBox.showinfo("Logout", "Thank you, Admin!", icon="info", fg="red")



if __name__ == '__main__':
    DisplayForm()
    tk.mainloop()
